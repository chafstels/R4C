import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from django.http import HttpResponse

from orders.models import Order
from .models import Robot
from .serializers import RobotSerializer
from django.db.models import Count
from django.db.models.functions import TruncWeek
from datetime import datetime, timedelta
from orders.tasks import send_email_task


def list_robots(request):
    robots = Robot.objects.all()
    serializer = RobotSerializer(robots, many=True)
    return JsonResponse(serializer.data, safe=False)


# Создание нового робота
@csrf_exempt
def create_robot(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            serializer = RobotSerializer(data=data)
            if serializer.is_valid():
                # Проверяем, есть ли заказы с указанной версией робота
                robot_serial = f'{data.get("model")}-{data.get("version")}'
                orders_with_robot = Order.objects.filter(robot_serial=robot_serial)
                print(orders_with_robot)
                print(robot_serial)

                if orders_with_robot:
                    # Отправляем уведомления на электронные адреса заказчиков
                    customer_emails = orders_with_robot.values_list(
                        "customer__email", flat=True
                    ).distinct()
                    print(customer_emails)
                    for email in customer_emails:
                        send_email_task.delay(
                            email=email,
                            robot_model=data["model"],
                            robot_serial=robot_serial,
                        )
                        print(email)

                serializer.save()
                return JsonResponse(
                    {"message": "The robot was successfully created"}, status=201
                )
            else:
                return JsonResponse({"error": serializer.errors}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Incorrect JSON format"}, status=400)
    else:
        return JsonResponse({"error": "The method is not supported"}, status=405)


@csrf_exempt
def update_robot(request, robot_id):
    try:
        robot = Robot.objects.get(pk=robot_id)
    except Robot.DoesNotExist:
        return JsonResponse({"error": "Robot not found"}, status=404)

    if request.method == "PUT":
        try:
            data = json.loads(request.body)
            serializer = RobotSerializer(robot, data=data)
            if serializer.is_valid():
                serializer.save()
                return JsonResponse(
                    {"message": "The robot has been successfully updated"}, status=200
                )
            else:
                return JsonResponse({"error": "Incorrect data"}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Incorrect JSON format"}, status=400)
    elif request.method == "PATCH":
        try:
            data = json.loads(request.body)
            serializer = RobotSerializer(
                robot, data=data, partial=True
            )  # Используем partial=True
            if serializer.is_valid():
                serializer.save()
                return JsonResponse(
                    {"message": "The robot has been successfully updated"}, status=200
                )
            else:
                return JsonResponse({"error": "Incorrect data"}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Incorrect JSON format"}, status=400)
    else:
        return JsonResponse({"error": "The method is not supported"}, status=405)


@csrf_exempt
def delete_robot(request, robot_id):
    try:
        robot = Robot.objects.get(pk=robot_id)
        robot.delete()
        return JsonResponse(
            {"message": "The robot was successfully removed"}, status=204
        )
    except Robot.DoesNotExist:
        return JsonResponse({"error": "Robot not found"}, status=404)


def generate_summary_excel(request):
    # Создаем новый Excel-документ
    workbook = openpyxl.Workbook()

    # Получаем уникальные модели роботов
    models = Robot.objects.values_list("model", flat=True).distinct()

    for model in models:
        sheet = workbook.create_sheet(title=model)

        sheet["A1"] = "Модель"
        sheet["B1"] = "Версия"
        sheet["C1"] = "Количество за неделю"

        # Фильтруем роботов только для текущей модели
        robots_for_model = (
            Robot.objects.filter(model=model).values("model", "version").distinct()
        )

        row = 2
        for robot in robots_for_model:
            sheet.cell(row=row, column=1, value=model)
            sheet.cell(row=row, column=2, value=robot["version"])
            sheet.cell(row=row, column=3, value=calculate_weekly_count(robot))
            row += 1

    # Удаляем страницу по умолчанию
    default_sheet = workbook.get_sheet_by_name("Sheet")
    workbook.remove(default_sheet)

    # Создаем HTTP-ответ для Excel-файла
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=robot_summary.xlsx"
    workbook.save(response)

    return response


def calculate_weekly_count(robot):
    current_datetime = datetime.now()
    start_of_week = current_datetime - timedelta(days=current_datetime.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    weekly_robots = (
        Robot.objects.filter(
            created__range=(start_of_week, end_of_week),
            model=robot["model"],
            version=robot["version"],
        )
        .values("model", "version")
        .annotate(total_count=Count("id"))
    )
    total_count = 0
    for robot_count in weekly_robots:
        total_count += robot_count["total_count"]

    return total_count
